import tkinter as tk
from gpt import clear_main_frame, show_students
import openpyxl
from openpyxl.styles import Alignment
from tkinter.filedialog import asksaveasfilename
from tkinter import messagebox

def export_to_excel(students):
    file_path = asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
        title="Lưu file Excel",
        initialfile="students.xlsx"
    )

    if not file_path:
        messagebox.showinfo("Hủy bỏ", "Bạn đã hủy lưu file Excel.")
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Students"

    headers = ["ID", "Name", "Age", "Major", "Class", "Grade"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num, student in enumerate(students, start=2):
        sheet.cell(row=row_num, column=1, value=student["id"])
        sheet.cell(row=row_num, column=2, value=student["name"])
        sheet.cell(row=row_num, column=3, value=student["age"])
        sheet.cell(row=row_num, column=4, value=student["major"])
        sheet.cell(row=row_num, column=5, value=student["Class"])
        sheet.cell(row=row_num, column=6, value=student["grade"])

    try:
        workbook.save(file_path)
        messagebox.showinfo("Thành công", f"Dữ liệu đã được lưu thành công tại:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Lỗi", f"Lỗi khi lưu file Excel: {e}")



def sort_students(students, main_frame):
    clear_main_frame(main_frame)

    tk.Label(main_frame, text="Sắp xếp sinh viên", font=("Arial", 16)).pack(pady=10)

    tk.Label(main_frame, text="Chọn tiêu chí sắp xếp:").pack(anchor="w", padx=10)
    criteria_var = tk.StringVar(value="name")
    criteria_options = ["name", "age", "grade"]

    for option in criteria_options:
        tk.Radiobutton(
            main_frame, text=option.capitalize(), variable=criteria_var, value=option
        ).pack(anchor="w", padx=20)

    tk.Label(main_frame, text="Chọn thứ tự sắp xếp:").pack(anchor="w", padx=10, pady=(10, 0))
    order_var = tk.StringVar(value="asc")
    tk.Radiobutton(
        main_frame, text="Tăng dần", variable=order_var, value="asc"
    ).pack(anchor="w", padx=20)
    tk.Radiobutton(
        main_frame, text="Giảm dần", variable=order_var, value="desc"
    ).pack(anchor="w", padx=20)

    def apply_sort():
        criteria = criteria_var.get()
        reverse = order_var.get() == "desc"

        if criteria == "name":
            students.sort(
                key=lambda student: (
                    student["name"].split()[-1].lower(),  # Họ
                    student["name"].split()[-2].lower(),  # Tên đệm
                    student["name"].split()[-3].lower()  # Tên
                ),
                reverse=reverse,
            )
        elif criteria == "age":
            students.sort(key=lambda student: int(student["age"]), reverse=reverse)
        elif criteria == "grade":
            students.sort(key=lambda student: float(student["grade"]), reverse=reverse)

        show_students(students, main_frame)

    tk.Button(main_frame, text="Sắp xếp", command=apply_sort).pack(pady=20)
